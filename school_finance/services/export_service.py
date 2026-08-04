"""Data export utilities for Excel format."""
import os
import datetime
from tkinter import filedialog

import openpyxl
from openpyxl.styles import Font, Alignment


def export_to_excel(data, headers, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            if isinstance(value, (int, float)):
                ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                ws.cell(row=row_idx, column=col_idx, value=str(value) or "")

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = max(len(str(h)) for h in headers) + 4

    wb.save(file_path)


def prompt_save_path(default_name, file_type_desc, extension):
    return filedialog.asksaveasfilename(
        title=f"Export as {file_type_desc}",
        defaultextension=extension,
        filetypes=[(file_type_desc, f"*{extension}")])


def export_students(students, base_dir):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = prompt_save_path(f"students_export_{timestamp}", "Excel files", ".xlsx")
    if not path:
        return None
    headers = ["ID", "Full Name", "Grade", "Stream", "Admission No.", "Status",
               "Waived", "Waiver Reason", "Waiver Date", "Remarks", "Balance (KES)"]
    export_to_excel(students, headers, path)
    return path


def export_payments(payments, base_dir):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = prompt_save_path(f"payments_export_{timestamp}", "Excel files", ".xlsx")
    if not path:
        return None
    headers = ["Receipt No.", "Student", "Grade", "Amount", "Method",
               "Date", "Term", "Received By"]
    export_to_excel(payments, headers, path)
    return path


def export_arrears(arrears, base_dir):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = prompt_save_path(f"arrears_export_{timestamp}", "Excel files", ".xlsx")
    if not path:
        return None
    headers = ["ID", "Full Name", "Grade", "Admission No.", "Balance (KES)"]
    data = [
        {
            "ID": a["id"],
            "Full Name": a["full_name"],
            "Grade": a["grade"],
            "Admission No.": a["admission_no"] or "",
            "Balance (KES)": a["balance"],
        }
        for a in arrears
    ]
    export_to_excel(data, headers, path)
    return path


def export_income_by_method(income_data, base_dir):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = prompt_save_path(f"income_by_method_{timestamp}", "Excel files", ".xlsx")
    if not path:
        return None
    headers = ["Method", "Total Amount (KES)"]
    data = [{"Method": k, "Total Amount (KES)": v} for k, v in income_data.items()]
    export_to_excel(data, headers, path)
    return path


def export_waived_students(students, base_dir):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = prompt_save_path(f"waived_students_{timestamp}", "Excel files", ".xlsx")
    if not path:
        return None
    headers = ["ID", "Full Name", "Grade", "Stream", "Admission No.",
                "Status", "Waiver Reason", "Waiver Date", "Remarks"]
    export_to_excel(students, headers, path)
    return path


def export_partial_waivers(waivers, base_dir):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = prompt_save_path(f"partial_waivers_{timestamp}", "Excel files", ".xlsx")
    if not path:
        return None
    headers = ["ID", "Student", "Grade", "Admission No.", "Term",
               "Gross Fee", "Waiver Amount", "Net Amount", "Reason",
               "Granted By", "Granted At", "Status"]
    export_to_excel(waivers, headers, path)
    return path
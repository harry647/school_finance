"""
One-off importer for the school's existing Excel balance sheets
(the Grade 7 / Grade 8 / Grade 9 style workbooks) and CSV files.

These sheets only ever recorded the *outstanding balance* per term, not a
full payment history. So each non-zero balance column found is imported as
a `charge` (money owed) for that student/term. From then on, real payments
recorded in the app reduce that balance normally.

Supported column header patterns (case-insensitive, matched by keyword):
  - "Name"                          -> student name (required)
  - "Admission No" / "Adm No" etc.  -> admission number (imported to the
                                       students table and used to avoid
                                       duplicate records)
  - "Grade" / "Class" / "Form"      -> the student's grade (used per row;
                                       falls back to the file name guess)
  - "Stream"                        -> student stream (imported if present)
  - "No." / "S/No"                  -> ignored (just a row counter)
  - "Balance 2025" / "2025 Balance" -> opening balance, year 2025
  - "Term I 2026" / "Term 1 Balance" etc. -> that term's balance
  - "Total ..." / "Total Balance"   -> ignored (recomputed by the app)
  - "Remarks" / "Comment" / "Notes" -> copied into the student's remarks

Import rules:
  * Students already in the system are matched by admission number first,
    then by (grade + name). They are skipped (not duplicated) and the
    missing admission number is back-filled when it is empty.
  * Identical imported charges (same student + term + description) are
    never inserted twice, so re-running an import is safe.
"""
import csv
import re
import os
import openpyxl

from db.database import get_connection
from models.student import add_student, list_students, update_student
from models.payment import add_charge
from models.term import get_or_create_term

TERM_WORD_MAP = {"1": "Term I", "i": "Term I",
                 "2": "Term II", "ii": "Term II",
                 "3": "Term III", "iii": "Term III"}

# Header keywords (matched after lower-casing + collapsing whitespace).
_NAME_HEADERS = {"name", "full name", "student name"}
_ADMISSION_HEADERS = {"admission", "admission no", "admission no.",
                      "admission number", "adm", "adm no", "adm no.",
                      "student no", "student no.", "student number",
                      "reg no", "reg no.", "reg", "roll no", "roll"}
_GRADE_HEADERS = {"grade", "class", "form", "grade/class"}

def _clean_text(value):
    """Collapse whitespace and strip a cell value, returning '' for None."""
    if value is None:
        return ""
    return str(value).strip()


def _normalize_grade(value):
    """Return a tidy 'Grade N' label from a grade cell when recognisable.

    Accepts 'Grade 7', 'Grade7', '7', 'Class 7', 'Form 7', '7 East', etc.
    Unknown values are returned unchanged.
    """
    s = _clean_text(value)
    if not s:
        return None
    low = s.lower()
    match = re.search(r"(\d{1,2})", s)
    if match and (low.startswith(("grade", "class", "form")) or low[0].isdigit()):
        return "Grade " + match.group(1)
    return s


def _to_amount(value):
    """Convert a cell to a float amount, tolerating strings like 'KSh 5,000'.

    Returns None when the cell has no usable number.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", s)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _guess_grade_from_filename(file_path):
    base = os.path.basename(file_path)
    match = re.search(r"grade\s*[_\- ]?(\d+)", base, re.IGNORECASE)
    if match:
        return f"Grade {match.group(1)}"
    return None

_STREAM_HEADERS = {"stream", "house"}
_REMARKS_HEADERS = {"remarks", "comment", "comments", "notes"}

def _find_header_row(sheet, max_scan=10):
    for row_idx in range(1, min(max_scan, sheet.max_row) + 1):
        values = [str(c.value).strip().lower() if c.value else ""
                  for c in sheet[row_idx]]
        if any(v == "name" for v in values):
            return row_idx, values
    raise ValueError("Could not find a header row containing a 'Name' column")


def _classify_columns(header_values, default_year):
    """Return dict: col_index(0-based) -> (kind, term_name, year).

    kind is one of: 'name' | 'admission' | 'grade' | 'stream' |
                    'remarks' | 'charge'
    """
    mapping = {}
    for idx, raw in enumerate(header_values):
        text = re.sub(r"\s+", " ", str(raw).strip().lower())
        if not text:
            continue
        if text in _NAME_HEADERS:
            mapping[idx] = ("name", None, None)
            continue
        if text in _ADMISSION_HEADERS:
            mapping[idx] = ("admission", None, None)
            continue
        if text in _GRADE_HEADERS:
            mapping[idx] = ("grade", None, None)
            continue
        if text in _STREAM_HEADERS:
            mapping[idx] = ("stream", None, None)
            continue
        if text in _REMARKS_HEADERS:
            mapping[idx] = ("remarks", None, None)
            continue
        if text.startswith("total") or re.fullmatch(r"(?:s/?no|no)\.?", text):
            continue  # row counters & total columns

        year_match = re.search(r"(20\d{2})", text)
        year = int(year_match.group(1)) if year_match else default_year

        term_match = re.search(r"term\s*([1-3]|i{1,3})\b", text)
        if term_match:
            term_name = TERM_WORD_MAP[term_match.group(1).lower()]
            mapping[idx] = ("charge", term_name, year)
        elif "balance" in text:
            # e.g. "Balance 2025" / "2025 Balance" with no explicit term
            # -> treat as an opening balance, not tied to a specific term
            mapping[idx] = ("charge", "Opening Balance", year)
    return mapping


def _find_header_row_csv(reader):
    for row in reader:
        if any(v.strip().lower() == "name" for v in row if v):
            return row
    raise ValueError("Could not find a 'Name' column in this CSV file")


def _classify_columns_csv(header_values, default_year):
    return _classify_columns(header_values, default_year)
def _import_rows(rows_iter, columns, file_grade, default_year,
                 existing_students):
    """Shared row processor for the Excel and CSV importers.

    ``rows_iter`` yields one list of cell values per data row.
    ``existing_students`` is the current student table loaded once.
    Returns a summary dict.
    """
    name_col = next((i for i, v in columns.items() if v[0] == "name"), None)
    if name_col is None:
        raise ValueError("No 'Name' column found in this file")
    admission_col = next((i for i, v in columns.items() if v[0] == "admission"),
                         None)
    grade_col = next((i for i, v in columns.items() if v[0] == "grade"), None)
    stream_col = next((i for i, v in columns.items() if v[0] == "stream"), None)
    remarks_col = next((i for i, v in columns.items() if v[0] == "remarks"),
                       None)

    # Build lookup maps ONCE (single pass over existing students) so the
    # import stays efficient even for large sheets.
    existing_by_admission = {}
    existing_by_name = {}
    for s in existing_students:
        d = dict(s)
        adm = _clean_text(d.get("admission_no"))
        if adm:
            existing_by_admission[adm.lower()] = d
        name_key = (str(d.get("grade", "")).strip().lower(),
                    _clean_text(d.get("full_name")).lower())
        existing_by_name.setdefault(name_key, d)

    # Memoise term lookups so each (year, term) hits the DB only once.
    term_cache = {}

    def _term_id(year, term_name):
        key = (year, term_name)
        if key not in term_cache:
            term_cache[key] = get_or_create_term(year, term_name)
        return term_cache[key]

    # Existing (student_id, term_id, description) charge keys — prevents
    # duplicating the same imported charge if an import is re-run.
    conn = get_connection()
    existing_charge_keys = {
        (r["student_id"], r["term_id"], _clean_text(r["description"]))
        for r in conn.execute(
            "SELECT student_id, term_id, description FROM charges"
        ).fetchall()
    }

    students_added = 0
    students_skipped = 0
    charges_added = 0
    admission_linked = 0
    duplicate_charges_skipped = 0
    rows_read = 0

    for values in rows_iter:
        values = list(values)
        if name_col >= len(values):
            continue
        name = _clean_text(values[name_col])
        if not name:
            continue
        rows_read += 1

        # Grade: an explicit column wins; otherwise use the file-level grade.
        row_grade = file_grade
        if grade_col is not None and grade_col < len(values):
            explicit_grade = _normalize_grade(values[grade_col])
            if explicit_grade:
                row_grade = explicit_grade

        admission_no = None
        if admission_col is not None and admission_col < len(values):
            admission_no = _clean_text(values[admission_col]) or None

        stream = None
        if stream_col is not None and stream_col < len(values):
            stream = _clean_text(values[stream_col]) or None

        remarks = None
        if remarks_col is not None and remarks_col < len(values):
            remarks = _clean_text(values[remarks_col]) or None

        # --- Find or create the student ---
        student = None
        if admission_no:
            student = existing_by_admission.get(admission_no.lower())
        if student is None:
            name_key = (row_grade.strip().lower(), name.lower())
            student = existing_by_name.get(name_key)

        if student is not None:
            student_id = student["id"]
            students_skipped += 1
            # Back-fill a missing admission number on an existing record.
            if admission_no and not _clean_text(student.get("admission_no")):
                update_student(student_id, admission_no=admission_no)
                student["admission_no"] = admission_no
                existing_by_admission[admission_no.lower()] = student
                admission_linked += 1
        else:
            student_id = add_student(
                full_name=name,
                grade=row_grade,
                admission_no=admission_no,
                stream=stream,
                remarks=remarks,
            )
            new_student = {
                "id": student_id,
                "full_name": name,
                "grade": row_grade,
                "admission_no": admission_no,
                "stream": stream,
                "remarks": remarks,
            }
            existing_by_name[(row_grade.strip().lower(), name.lower())] = new_student
            if admission_no:
                existing_by_admission[admission_no.lower()] = new_student
            students_added += 1

        # --- Import charge columns ---
        for idx, (kind, term_name, year) in columns.items():
            if kind != "charge" or idx >= len(values):
                continue
            amount = _to_amount(values[idx])
            if amount is None or amount <= 0:
                continue
            term_id = _term_id(year, term_name)
            description = f"Imported balance - {term_name} {year}"
            key = (student_id, term_id, description)
            if key in existing_charge_keys:
                duplicate_charges_skipped += 1
                continue
            existing_charge_keys.add(key)
            add_charge(student_id, amount, term_id=term_id,
                       description=description)
            charges_added += 1

    return {
        "rows_read": rows_read,
        "students_added": students_added,
        "students_skipped": students_skipped,
        "charges_added": charges_added,
        "admission_linked": admission_linked,
        "duplicate_charges_skipped": duplicate_charges_skipped,
    }

def import_balance_sheet(file_path, default_grade=None, default_year=2026):
    """
    Import one legacy balance-sheet workbook.

    Returns a summary dict with keys: grade, rows_read, students_added,
    students_skipped, charges_added, admission_linked,
    duplicate_charges_skipped.
    """
    grade = _normalize_grade(default_grade) or \
        _guess_grade_from_filename(file_path) or "Unknown Grade"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.worksheets[0]

    header_row, header_values = _find_header_row(sheet)
    columns = _classify_columns(header_values, default_year)

    rows_iter = (
        [c.value for c in row]
        for row in sheet.iter_rows(min_row=header_row + 1,
                                   max_row=sheet.max_row)
    )

    result = _import_rows(rows_iter, columns, grade, default_year,
                          list_students())
    result["grade"] = grade
    return result


def import_csv_balance_sheet(file_path, default_grade=None, default_year=2026):
    """Import one legacy balance-sheet CSV file (same columns as the Excel)."""
    grade = _normalize_grade(default_grade) or \
        _guess_grade_from_filename(file_path) or "Unknown Grade"

    with open(file_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header_values = _find_header_row_csv(reader)
        columns = _classify_columns_csv(header_values, default_year)
        result = _import_rows(reader, columns, grade, default_year,
                              list_students())
    result["grade"] = grade
    return result


def import_balance_sheet_folder(folder_path, default_year=2026):
    """Import every balance sheet (.xlsx/.xlsm/.csv) in a folder.

    Returns an aggregate summary with a per-file breakdown so the UI can
    show exactly what each file contributed.
    """
    if not os.path.isdir(folder_path):
        raise ValueError(f"Not a folder: {folder_path}")

    excel_files = sorted(
        f for f in os.listdir(folder_path)
        if f.lower().endswith((".xlsx", ".xlsm"))
    )
    csv_files = sorted(
        f for f in os.listdir(folder_path)
        if f.lower().endswith(".csv")
    )
    if not excel_files and not csv_files:
        raise ValueError("No Excel (.xlsx/.xlsm) or CSV files found in folder")

    totals = {
        "files_processed": 0,
        "rows_read": 0,
        "students_added": 0,
        "students_skipped": 0,
        "charges_added": 0,
        "admission_linked": 0,
        "duplicate_charges_skipped": 0,
    }
    file_results = []

    for file_name in excel_files + csv_files:
        path = os.path.join(folder_path, file_name)
        file_grade = _guess_grade_from_filename(path)
        try:
            if file_name.lower().endswith(".csv"):
                result = import_csv_balance_sheet(
                    path, default_grade=file_grade, default_year=default_year)
            else:
                result = import_balance_sheet(
                    path, default_grade=file_grade, default_year=default_year)
        except Exception as exc:
            file_results.append({"file": file_name, "error": str(exc)})
            continue

        file_results.append({"file": file_name, **result})
        totals["files_processed"] += 1
        for key in ("rows_read", "students_added", "students_skipped",
                    "charges_added", "admission_linked",
                    "duplicate_charges_skipped"):
            totals[key] += result.get(key, 0)

    totals["files"] = file_results
    return totals


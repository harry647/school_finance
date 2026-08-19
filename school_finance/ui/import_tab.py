import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import logging

from models.user import log_action
from services.import_service import (
    import_balance_sheet,
    import_balance_sheet_folder,
)
from ui.constants import FONT_MUTED, MUTED_FG, PAD_LG, PAD_MD, PAD_SM, PAD_XS, SUCCESS

logger = logging.getLogger("school_finance")


class ImportTab(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, padding=PAD_MD)
        self.app = app
        self._build_ui()
        self.refresh()

    def refresh(self):
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")
        self.progress_var.set("")
        self.progress_bar.stop()
        self.import_btn.config(state="normal")
        self.folder_btn.config(state="normal")

    def _build_ui(self):
        box = ttk.LabelFrame(self, text="Import Legacy Excel Balance Sheet",
                                     padding=PAD_MD)
        box.pack(fill="x")

        ttk.Label(
            box,
            text=("Import existing balance-sheet workbook(s) (e.g. the "
                  "Grade 7 / Grade 8 / Grade 9 Balance Sheet files). Each "
                  "student's outstanding balance per term is imported as a "
                  "charge. Students already in the system are matched by "
                  "admission number or (grade + name) and skipped — not "
                  "duplicated."),
            wraplength=560, foreground=MUTED_FG,
        ).pack(anchor="w", pady=(0, PAD_MD))

        ttk.Label(
            box,
            text=("All columns are captured: 'Name' (required), "
                  "'Admission No', 'Grade', 'Stream' and 'Remarks' are saved "
                  "on the student record; balance columns like 'Term I 2026', "
                  "'Term II 2026' and 'Balance 2025' become charges. "
                  "Download the template below for the exact format."),
            wraplength=560, foreground=SUCCESS,
        ).pack(anchor="w", pady=(0, PAD_MD))

        btn_row = ttk.Frame(box)
        btn_row.pack(fill="x", pady=(0, PAD_MD))

        self.import_btn = ttk.Button(btn_row, text="Choose Excel File & Import",
                        command=self._choose_and_import)
        self.import_btn.pack(side="left", padx=(0, PAD_SM))
        self.folder_btn = ttk.Button(btn_row, text="Choose Folder & Import All",
                        command=self._choose_folder_and_import)
        self.folder_btn.pack(side="left", padx=(0, PAD_SM))
        ttk.Button(btn_row, text="Download Import Template",
                        command=self._download_template).pack(side="left")

        row = ttk.Frame(box)
        row.pack(fill="x", pady=PAD_XS)
        ttk.Label(row, text="Grade for this file (optional — auto-detected "
                             "from filename if left blank):").pack(side="left")
        self.grade_var = tk.StringVar()
        ttk.Entry(row, textvariable=self.grade_var, width=15).pack(side="left",
                                                                     padx=PAD_SM)

        self.progress_var = tk.StringVar(value="")
        ttk.Label(box, textvariable=self.progress_var,
                       foreground=MUTED_FG).pack(anchor="w", pady=(PAD_SM, 0))

        self.progress_bar = ttk.Progressbar(box, mode="indeterminate")
        self.progress_bar.pack(fill="x", pady=(PAD_XS, 0))

        self.log_text = tk.Text(self, height=14, state="disabled",
                                        wrap="word")
        self.log_text.pack(fill="both", expand=True, pady=(PAD_MD, 0))

    def _log(self, message):
        self.log_text.config(state="normal")
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def _download_template(self):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Import Template"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            headers = ["Name", "Admission No", "Grade", "Stream", "Term I 2026",
                       "Term II 2026", "Term III 2026", "Balance 2025", "Remarks"]
            ws.append(headers)

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            sample_data = [
                ["John Ochieng", "GVSS/2026/001", "Grade 7", "East",
                 15000, 0, 0, 5000, "Parent promised payment"],
                ["Mary Wanjiku", "GVSS/2026/002", "Grade 7", "East",
                 0, 18000, 0, 0, ""],
                ["Peter Njoroge", "GVSS/2026/003", "Grade 8", "West",
                 25000, 25000, 25000, 10000, ""],
            ]
            for row_data in sample_data:
                ws.append(row_data)

            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 14
            ws.column_dimensions["F"].width = 14
            ws.column_dimensions["G"].width = 14
            ws.column_dimensions["H"].width = 14
            ws.column_dimensions["I"].width = 28

            notes = wb.create_sheet("Format Notes")
            notes.column_dimensions["A"].width = 80
            notes.append(["IMPORT TEMPLATE FORMAT NOTES"])
            notes["A1"].font = Font(bold=True, size=14)
            notes.append([])
            notes.append(["REQUIRED COLUMNS:"])
            notes["A3"].font = Font(bold=True)
            notes.append(["• Name - Student full name (required)"])
            notes.append([])
            notes.append(["STUDENT DETAIL COLUMNS (all optional):"])
            notes["A5"].font = Font(bold=True)
            notes.append(["• Admission No - Student admission number"])
            notes.append(["• Grade - e.g. Grade 7 (or leave blank; detected "
                          "from the file name)"])
            notes.append(["• Stream - e.g. East / West (optional)"])
            notes.append(["• Remarks - Any notes about the student"])
            notes.append([])
            notes.append(["BALANCE COLUMNS (optional but recommended):"])
            notes["A10"].font = Font(bold=True)
            notes.append(["• Use format: Term I 2026, Term II 2026, "
                          "Term III 2026, etc."])
            notes.append(["• Or use: Balance 2025, Balance 2026 for opening "
                          "balances"])
            notes.append(["• Any column with a year (2024, 2025, 2026, etc.) "
                          "is detected"])
            notes.append([])
            notes.append(["IMPORT RULES:"])
            notes["A15"].font = Font(bold=True)
            notes.append(["• Students are matched by Admission No first, then "
                          "(Grade + Name)"])
            notes.append(["• Existing students are skipped, not duplicated"])
            notes.append(["• Only positive numeric values are imported as "
                          "charges"])
            notes.append(["• Re-importing the same file is safe — identical "
                          "charges are skipped"])
            notes.append(["• Column headers are case-insensitive"])
            notes.append(["• Save as .xlsx, .xlsm, or .csv before importing"])

            default_dir = self.app.receipts_dir
            path = os.path.join(default_dir, "import_template.xlsx")
            wb.save(path)

            messagebox.showinfo("Template Downloaded",
                                f"Template saved to:\n{path}\n\n"
                                "Fill in your student data and import it using the button above.")
            logger.info("Import template downloaded to: %s", path)
            self._log(f"Template downloaded to: {path}")
        except Exception as e:
            logger.error("Failed to create template: %s", e, exc_info=True)
            messagebox.showerror("Error", f"Failed to create template: {e}")

    def _choose_and_import(self):
        file_path = filedialog.askopenfilename(
            title="Select balance sheet Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm"),
                        ("CSV files", "*.csv")])
        if not file_path:
            return
        self.import_btn.config(state="disabled")
        self.progress_var.set("Importing...")
        self.progress_bar.start(10)
        self.update_idletasks()
        try:
            result = import_balance_sheet(
                file_path, default_grade=self.grade_var.get().strip()
                or None)
        except Exception as e:
            logger.error("Import failed for %s: %s", file_path, e, exc_info=True)
            messagebox.showerror("Import failed", str(e))
            self._log(f"FAILED: {file_path} -> {e}")
        else:
            msg = (f"Imported '{file_path}'\n"
                   f"  Grade: {result['grade']}\n"
                   f"  New students added: {result['students_added']}\n"
                   f"  Existing students skipped: {result['students_skipped']}\n"
                   f"  Charges recorded: {result['charges_added']}\n"
                   f"  Admission numbers linked: {result.get('admission_linked', 0)}\n"
                   f"  Duplicate charges skipped: {result.get('duplicate_charges_skipped', 0)}")
            self._log(msg)
            log_action(self.app.current_username, "import_excel",
                       msg.replace("\n", " | "))
            messagebox.showinfo("Import complete",
                                    f"Added {result['students_added']} students, "
                                    f"recorded {result['charges_added']} charges.")
            self.app.refresh_all()
        finally:
            self.progress_bar.stop()
            self.progress_var.set("")
            self.import_btn.config(state="normal")
            self.folder_btn.config(state="normal")

    def _choose_folder_and_import(self):
        folder_path = filedialog.askdirectory(
            title="Select folder containing balance sheet files")
        if not folder_path:
            return
        self.folder_btn.config(state="disabled")
        self.import_btn.config(state="disabled")
        self.progress_var.set("Importing all files in folder...")
        self.progress_bar.start(10)
        self.update_idletasks()
        try:
            result = import_balance_sheet_folder(folder_path)
        except Exception as e:
            logger.error("Folder import failed for %s: %s", folder_path, e,
                         exc_info=True)
            messagebox.showerror("Folder import failed", str(e))
            self._log(f"FAILED: {folder_path} -> {e}")
        else:
            failed = [f["file"] for f in result["files"] if "error" in f]
            lines = [f"Imported folder: {folder_path}",
                     f"  Files processed: {result['files_processed']}",
                     f"  New students added: {result['students_added']}",
                     f"  Existing students skipped: {result['students_skipped']}",
                     f"  Charges recorded: {result['charges_added']}",
                     f"  Admission numbers linked: {result['admission_linked']}",
                     f"  Duplicate charges skipped: {result['duplicate_charges_skipped']}"]
            for fr in result["files"]:
                if "error" in fr:
                    lines.append(f"  [FAILED] {fr['file']} -> {fr['error']}")
                else:
                    lines.append(f"  [{fr.get('grade', '?')}] {fr['file']}: "
                                 f"+{fr['students_added']} students, "
                                 f"{fr['charges_added']} charges")
            msg = "\n".join(lines)
            self._log(msg)
            log_action(self.app.current_username, "import_excel_folder",
                       msg.replace("\n", " | "))
            summary = (f"Imported {result['students_added']} new students and "
                       f"{result['charges_added']} charges across "
                       f"{result['files_processed']} file(s).")
            if failed:
                summary += f"\n\n{failed}"
            messagebox.showinfo("Folder import complete", summary)
            self.app.refresh_all()
        finally:
            self.progress_bar.stop()
            self.progress_var.set("")
            self.folder_btn.config(state="normal")
            self.import_btn.config(state="normal")
